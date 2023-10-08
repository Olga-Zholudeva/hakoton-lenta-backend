from django.http import HttpResponse
from django.db.models import Max
from django_filters.rest_framework import DjangoFilterBackend

from rest_framework import viewsets, mixins
from rest_framework.decorators import action
from rest_framework.permissions import SAFE_METHODS
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_workbook

from api.serializers import (SalesSerializer, SalesPostSerializer,
                             StoreSerializer, SkuSerializer,
                             ForecastSerializer, ForecastPostSerializer,
                             SalesDiffSerializer, NewForecastSerializer)
from api.filters import SalesFilter, ForecastFilter, SalesDiffFilter
from products.models import Sku, SalesFact, Store, Forecast, SalesDiff


class StoreViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet
):
    '''Обработчик для магазинов'''
    queryset = Store.objects.all()
    serializer_class = StoreSerializer


class SkuViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet
):
    '''Обработчик для товаров'''
    queryset = Sku.objects.all()
    serializer_class = SkuSerializer


class SalesViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet
):
    '''Обработчик для фактических продаж'''
    queryset = SalesFact.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = SalesFilter
    filterset_fields = ['city', 'store', 'sku', 'group',
                        'category', 'subcategory', 'date_from', 'date_to']

    def get_serializer_class(self):
        if self.request.method in SAFE_METHODS:
            return SalesSerializer
        return SalesPostSerializer


class ForecastViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet
):
    '''Обработчик для прогноза'''
    queryset = Forecast.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = ForecastFilter
    filterset_fields = ['city', 'store', 'sku', 'group',
                        'category', 'subcategory', 'date_from', 'date_to']

    def get_serializer_class(self):
        if self.request.method in SAFE_METHODS:
            return ForecastSerializer
        return ForecastPostSerializer
    
    @action(methods=["GET",], detail=False, serializer_class=ForecastSerializer)
    def dowload(self, request, *args, **kwargs):
        last_forecast = Forecast.objects.aggregate(
            Max('forecast_date')
        )['forecast_date__max']
        queryset = Forecast.objects.filter(
            forecast_date=last_forecast
        )
        query_params = request.query_params
        if 'city' and 'store' and 'group' and 'category' and 'subcategory' and 'sku' and 'date_from' and 'date_to' in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
                st_sku_date__st_id=query_params['store'],
                st_sku_date__pr_sku_id__pr_group_id=query_params['group'],
                st_sku_date__pr_sku_id__pr_cat_id=query_params['category'],
                st_sku_date__pr_sku_id__pr_subcat_id=query_params['subcategory'],
                st_sku_date__pr_sku_id=query_params['sku'],
                st_sku_date__date=query_params['date_from'],
                date_to=query_params['date_to']
            )
        elif 'city' and 'store' and 'group' and 'category' and 'subcategory' and 'sku' and 'date_from' in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
                st_sku_date__st_id=query_params['store'],
                st_sku_date__pr_sku_id__pr_group_id=query_params['group'],
                st_sku_date__pr_sku_id__pr_cat_id=query_params['category'],
                st_sku_date__pr_sku_id__pr_subcat_id=query_params['subcategory'],
                st_sku_date__pr_sku_id=query_params['sku'],
                st_sku_date__date=query_params['date_from'],
            )
        elif 'city' and 'store' and 'group' and 'category' and 'subcategory' and 'sku'in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
                st_sku_date__st_id=query_params['store'],
                st_sku_date__pr_sku_id__pr_group_id=query_params['group'],
                st_sku_date__pr_sku_id__pr_cat_id=query_params['category'],
                st_sku_date__pr_sku_id__pr_subcat_id=query_params['subcategory'],
                st_sku_date__pr_sku_id=query_params['sku'],
            )
        elif 'city' and 'store' and 'group' and 'category' and 'subcategory' in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
                st_sku_date__st_id=query_params['store'],
                st_sku_date__pr_sku_id__pr_group_id=query_params['group'],
                st_sku_date__pr_sku_id__pr_cat_id=query_params['category'],
                st_sku_date__pr_sku_id__pr_subcat_id=query_params['subcategory'],
            )
        elif 'city' and 'store' and 'group' and 'category' in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
                st_sku_date__st_id=query_params['store'],
                st_sku_date__pr_sku_id__pr_group_id=query_params['group'],
                st_sku_date__pr_sku_id__pr_cat_id=query_params['category'],
            )
        elif 'city' and 'store' and 'group' in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
                st_sku_date__st_id=query_params['store'],
                st_sku_date__pr_sku_id__pr_group_id=query_params['group'],
            )
        elif 'city' and 'store' in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
                st_sku_date__st_id=query_params['store'],
            )
        elif 'city' in query_params:
            queryset = Forecast.objects.filter(
                st_sku_date__st_id__st_city_id=query_params['city'],
            ) 
        workbook = Workbook()
        worksheet = workbook.active

        headers = ['store', 'group', 'category', 'subcategory', 'sku', 'date', 'sales_units']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header

        for row_num, forecast in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=str(forecast.st_sku_date.st_id))
            worksheet.cell(row=row_num, column=2, value=forecast.st_sku_date.pr_sku_id.pr_group_id)
            worksheet.cell(row=row_num, column=3, value=forecast.st_sku_date.pr_sku_id.pr_cat_id)
            worksheet.cell(row=row_num, column=4, value=forecast.st_sku_date.pr_sku_id.pr_subcat_id)
            worksheet.cell(row=row_num, column=5, value=str(forecast.st_sku_date.pr_sku_id))
            worksheet.cell(row=row_num, column=6, value=str(forecast.st_sku_date.date))
            worksheet.cell(row=row_num, column=7, value=forecast.sales_units)

        filename = 'forecast_data.xlsx'
        workbook.save(filename)
        with open(filename, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="{0}"'.format(filename)
            return response


class SalesDiffViewSet(
    mixins.ListModelMixin,
    viewsets.GenericViewSet
):
    '''Обработчик для качества'''
    queryset = SalesDiff.objects.all()
    serializer_class = SalesDiffSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = SalesDiffFilter
    filterset_fields = ['city', 'store', 'sku', 'group',
                        'category', 'subcategory', 'date_from', 'date_to']

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        if 'format' in request.GET and request.GET['format'].lower() == 'excel':
            serializer = self.get_serializer(queryset, many=True)
            workbook = Workbook()
            worksheet = workbook.active

            headers = serializer.fields.keys()
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                worksheet.cell(row=1, column=col_num, value=header)

            for row_num, row_data in enumerate(serializer.data, 2):
                for col_num, field_name in enumerate(headers, 1):
                    column_letter = get_column_letter(col_num)
                    cell_value = row_data[field_name]
                    worksheet.cell(row=row_num, column=col_num,
                                   value=cell_value)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="forecast.xlsx"'
            response.write(save_workbook(workbook))
            return response

        else:
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)


class NewForecastViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    queryset = Forecast.objects.all()
    serializer_class = NewForecastSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = ForecastFilter
    filterset_fields = ['city', 'store', 'sku', 'group', 'category',
                        'subcategory', 'date_from', 'date_to', 'forecast_date']
